import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = r"D:/Hwan/Documents/Web/edu-kit"
def raw(name):
    with open(os.path.join(ROOT, "data/raw", name), encoding="utf-8") as f:
        return json.load(f)

kits = raw("kits.json"); items = raw("items.json"); stage_meta = raw("stage_meta.json")

FONT = "맑은 고딕"
HEAD_FILL = PatternFill("solid", fgColor="1F7AF0")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name=FONT, size=10)
thin = Side(style="thin", color="D9E2EC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active; ws.title = "읽어주세요"
ws.sheet_properties.tabColor = "1F7AF0"
guide = [
    ("수업꾸러미 콘텐츠 시트", 16, True, "1F7AF0"),
    ("", 10, False, None),
    ("이 통합문서를 Google Drive에 올린 뒤 'Google Sheets로 열기'로 변환하세요.", 11, False, None),
    ("탭은 정확히 kits / items / stage_meta 세 개입니다(이름·헤더 그대로 유지).", 11, False, None),
    ("", 10, False, None),
    ("■ 발행 방법 (요약)", 12, True, "1F7AF0"),
    ("1) 시트에 내용을 입력/수정합니다.", 11, False, None),
    ("2) 확장 프로그램 > Apps Script 에 gas/publish.gs 를 붙여넣고,", 11, False, None),
    ("   스크립트 속성 GITHUB_TOKEN(레포 Contents 쓰기 권한 PAT)을 등록합니다.", 11, False, None),
    ("3) 시트에 버튼(그림)을 넣고 스크립트 '발행'을 연결합니다.", 11, False, None),
    ("4) [발행] 클릭 -> GitHub 커밋 -> Cloudflare Pages 자동 빌드/배포(수 분).", 11, False, None),
    ("   잘못된 값이 있으면 빌드가 실패해 배포되지 않습니다(안전망).", 11, False, None),
    ("", 10, False, None),
    ("■ 꼭 알아둘 점", 12, True, "1F7AF0"),
    ("- kits의 id 칸은 비워 두면 발행 시 자동 부여(base31 4자)되고 시트에 고정됩니다.", 11, False, None),
    ("- 공개하려면 published = TRUE. 준비 중이면 FALSE.", 11, False, None),
    ("- 흐름형(flow) 항목: title=핵심용어, video_title=영상제목, video_desc=영상설명.", 11, False, None),
    ("- 본문(body)은 제한 마크다운만: ### 제목, **굵게**, 표, - 목록, > 인용, :::warn / :::info.", 11, False, None),
    ("- 이미지/영상은 출처/라이선스(image_source/license, video_license)를 채워주세요.", 11, False, None),
    ("- 조회수/좋아요는 시트에 없습니다(런타임 집계).", 11, False, None),
    ("", 10, False, None),
    ("자세한 가이드: 레포의 docs/SHEET_TEMPLATE.md , docs/DEPLOY.md", 10, False, None),
]
for i, (text, size, bold, color) in enumerate(guide, start=1):
    c = ws.cell(row=i, column=1, value=text)
    c.font = Font(name=FONT, size=size, bold=bold, color=(color or "1F2937"))
ws.column_dimensions["A"].width = 95
ws.sheet_view.showGridLines = False

def make_sheet(name, headers, rows, validations, widths):
    ws = wb.create_sheet(name)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HEAD_FONT; c.fill = HEAD_FILL
        c.alignment = Alignment(vertical="center"); c.border = BORDER
    for i, row in enumerate(rows, start=2):
        for j, h in enumerate(headers, start=1):
            v = row.get(h, "")
            c = ws.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.alignment = Alignment(vertical="top", wrap_text=(h == "body"))
            c.border = BORDER
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for col_letter, options in validations.items():
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(options), allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("%s2:%s1000" % (col_letter, col_letter))
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

kit_headers = ["id","title","grade","sem","subject","unit","unit_no","flow","sort_order","published"]
make_sheet("kits", kit_headers, kits,
    {"C": ["3","4","5","6"], "D": ["1학기","2학기"], "E": ["사회","과학"],
     "H": ["activity","flow"], "J": ["TRUE","FALSE"]},
    {"A":8,"B":24,"C":7,"D":8,"E":8,"F":26,"G":9,"H":10,"I":10,"J":11})

item_headers = ["kit_id","item_key","stage","type","title","description","sort_order",
    "core_idea","core_question","concepts","standard_code","standard_text",
    "video_url","start_sec","end_sec","video_title","video_desc","video_license",
    "caption","image_url","image_label","image_sub","image_source","image_license","body"]
make_sheet("items", item_headers, items,
    {"C": ["단원안내","생각열기","탐구하기","확장하기","도입","전개","정리"],
     "D": ["intro","video","image","text"]},
    {"A":8,"B":9,"C":10,"D":7,"E":22,"F":24,"G":9,"H":28,"I":28,"J":26,"K":12,"L":30,
     "M":34,"N":9,"O":9,"P":26,"Q":34,"R":12,"S":26,"T":18,"U":16,"V":22,"W":16,"X":14,"Y":40})

sm_headers = ["kit_id","stage","question","sort_order"]
make_sheet("stage_meta", sm_headers, stage_meta,
    {"B": ["단원안내","생각열기","탐구하기","확장하기","도입","전개","정리"]},
    {"A":8,"B":10,"C":44,"D":10})

out = os.path.join(ROOT, "docs", "수업꾸러미_시트.xlsx")
wb.save(out)
print("saved:", out)
print("sheets:", wb.sheetnames)
print("rows kits/items/stage_meta:", len(kits), len(items), len(stage_meta))
